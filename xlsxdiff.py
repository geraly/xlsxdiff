import argparse

import openpyxl as px


# sheet1: openpyxl.worksheet.worksheet.Worksheet
# sheet2: openpyxl.worksheet.worksheet.Worksheet
# target_range : string
#         ex. 'A1:B2'
def xlsxdiff(sheet1, sheet2, target_range: str):
    target1 = sheet1[target_range]
    target2 = sheet2[target_range]

    for row1, row2 in zip(target1, target2):
        for cell1, cell2 in zip(row1, row2):
            if cell1.value != cell2.value:
                print(
                    f"{cell1.coordinate} {'None' if cell1.value is None else cell1.value}"
                )
                print(
                    f"{cell2.coordinate} {'None' if cell2.value is None else cell2.value}"
                )


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="diff xlsx")
    parser.add_argument("--sheetname", help="Specify sheets to diff")
    parser.add_argument("filename1")
    parser.add_argument("filename2")
    parser.add_argument("range", help="ex. A1:B2")
    args = parser.parse_args()

    workbook1 = px.load_workbook(args.filename1, data_only=True)
    workbook2 = px.load_workbook(args.filename2, data_only=True)

    if args.sheetname is None:
        sheetnames1 = workbook1.sheetnames
        sheetnames2 = workbook2.sheetnames
    else:
        sheetnames1 = [args.sheetname]
        sheetnames2 = [args.sheetname]

    for sheetname in sheetnames2:
        if sheetname not in sheetnames1:
            print(f"{sheetname} is not in {args.filename1}.")
            continue
        print(f"diff {args.filename1} {args.filename2} {sheetname}")
        xlsxdiff(workbook1[sheetname], workbook2[sheetname], args.range)
        sheetnames1.remove(sheetname)
        sheetnames2.remove(sheetname)

    for sheetname in sheetnames1:
        if sheetname not in sheetnames2:
            print(f"{sheetname} is not in {args.filename2}.")
            continue
        print(f"diff {args.filename1} {args.filename2} {sheetname}")
        xlsxdiff(workbook1[sheetname], workbook2[sheetname], args.range)

    workbook1.close()
    workbook2.close()
